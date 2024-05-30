import classes
import openpyxl


def print_lecture(courses,professors):
    file = openpyxl.Workbook()
    sheet = file.create_sheet(title='Courses', index=0)
    i = 1
    j = 0
    for course in courses:
        for day in course.scheduled:
            sheet.cell(row=day.pos, column=i).value = day.lecture
            j = j + 1
        i = i + 1
        j = 0

    i = 1
    j = 2
    sheet = file.create_sheet(title='Availability', index=1)
    for prof in professors:
        sheet.cell(row=i, column=1).value = professors[prof].name
        for time in professors[prof].available_days:
            sheet.cell(row=i, column=j).value = time
            j = j + 1
        i = i + 1
        j = 2
    file.save(filename = 'Output.xlsx')

def print_prof_availability(professors):
    file = openpyxl.Workbook('Output.xlsx')







def find_sessions(lecture,days,professors,to_schedule,pos,orig):
    current_day = days.head
    while current_day is not None:
        if abs(current_day.value.pos - pos) == 1 and current_day.value.pos != orig:
            if current_day.value.date in professors[lecture.professor].available_days:
                to_schedule = to_schedule - 1
                if to_schedule != 0:

                    found = find_sessions(lecture,days,professors,to_schedule,current_day.value.pos,pos)
                    if found is None:
                        to_schedule = to_schedule + 1
                    else:
                        return found
                else:
                    return current_day.value.pos
        current_day = current_day.next
    return None


def schedule_classes(courses, professors,sessions):
    lever = len(courses)
    i = 0
    scheduled_days = [0] * sessions
    while lever != 0:
        if courses[i].days.head is not None:
            lever = len(courses)
            courses[i].days.sorting_days(scheduled_days)
            scheduled_days = [0] * sessions
            day = courses[i].days.pop()
            courses[i].lectures.sorting_lectures(professors, courses[i].heatmaps,day.pos)
            lecture = courses[i].lectures.head
            lecture_prev = None
            switch = True
            if day.pos == 16 and i == 0:
                print("hello")
            while lecture is not None and switch is True:
                if day.date in professors[lecture.value.professor].available_days:
                    to_schedule = lecture.value.length
                    if to_schedule == 1:
                        switch = False
                        day.lecture = lecture.value.name
                        courses[i].scheduled.append(day)
                        courses[i].heating(day.pos, lecture.value.category)
                        scheduled_days[day.pos - 1] = scheduled_days[day.pos - 1] + 1
                        professors[lecture.value.professor].available_days.remove(day.date)
                        professors[lecture.value.professor].remove_occurrence(lecture.value.length)
                        if lecture_prev is None:
                            courses[i].lectures.head = lecture.next
                        else:
                            lecture_prev.next = lecture.next
                    else:
                        found = find_sessions(lecture.value,courses[i].days,professors,to_schedule - 1,day.pos,0)
                        if found is not None:
                            switch = False
                            day.lecture = lecture.value.name
                            courses[i].scheduled.append(day)
                            courses[i].heating(day.pos,lecture.value.category)
                            scheduled_days[day.pos - 1] = scheduled_days[day.pos - 1] + 1
                            professors[lecture.value.professor].available_days.remove(day.date)
                            professors[lecture.value.professor].remove_occurrence(lecture.value.length)
                            if lecture_prev is None:
                                courses[i].lectures.head = lecture.next
                            else:
                                lecture_prev.next = lecture.next
                            temp_day = courses[i].days.head
                            day_prev = None
                            x = 0
                            y = 0
                            if day.pos < found:
                                x = day.pos
                                y = found
                            else:
                                x = found
                                y = day.pos
                            while temp_day is not None:
                                if x <= temp_day.value.pos <= y:
                                    professors[lecture.value.professor].available_days.remove(temp_day.value.date)
                                    professors[lecture.value.professor].remove_occurrence(lecture.value.length)
                                    temp_day.value.lecture = lecture.value.name
                                    courses[i].heating(day.pos, lecture.value.category)
                                    courses[i].scheduled.append(temp_day.value)
                                    scheduled_days[temp_day.value.pos - 1] = scheduled_days[temp_day.value.pos - 1] + 1
                                    if day_prev is None:
                                        courses[i].days.head = temp_day.next
                                    else:
                                        day_prev.next = temp_day.next
                                        temp_day = day_prev
                                day_prev = temp_day
                                temp_day = temp_day.next
                                if lecture_prev is None:
                                    courses[i].lectures.head = lecture.next
                                else:
                                    lecture_prev.next = lecture.next

                lecture_prev = lecture
                lecture = lecture.next

            if switch is True:
                courses[i].not_scheduled.append(day)
        else:
            lever = lever - 1

        if i is len(courses) - 1:
            i = 0
        else:
            i = i + 1

    print_lecture(courses,professors)











