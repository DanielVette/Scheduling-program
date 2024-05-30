from typing import List, Any
import classes as c

import openpyxl

courses = []
course_list = None
wish_sheet = None
availability = None


def preparing_database(filename):
    data_file = openpyxl.load_workbook(filename)
    sheets = data_file.sheetnames

    # loading all Lecture data
    global course_list
    course_list = data_file[sheets[0]]

    global wish_sheet
    wish_sheet = data_file[sheets[1]]

    global availability
    availability = data_file[sheets[2]]


# pars_profs helper function to find column with the professors availability
def find_prof(name):
    i = 2

    global availability
    while availability.cell(row=1, column=i).value is not None:
        if availability.cell(row=1, column=i).value is name:
            return i
        i = i + 1

    print("cant find professor",name)
    exit()


# create an array with all professors
def pars_profs():

    i = 2
    lectures = []
    times = []
    professors = {}
    global course_list
    while course_list.cell(row=i, column=2).value is not None:
        name = course_list.cell(row=i, column=2).value
        while name is course_list.cell(row=i, column=2).value:
            lectures.append(course_list.cell(row=i, column=1).value)
            i = i + 1

        j = 2
        col = find_prof(name)
        if col is not None:
            global availability
            while availability.cell(row=j, column=1).value is not None:
                if availability.cell(row=j, column=col).value is not None:
                    times.append(availability.cell(row=j, column=1).value)
                j = j + 1

        professors[name] = c.Professor(name, lectures, times)
        times = []
    return professors


# create a sorted link list of days needing to scheduled for each course
def pars_days():
    i = 2
    j = 2
    days = c.LinkedList()

    while availability.cell(row=i,column=1).value is not None:
        available_profs = 0
        while availability.cell(row=1,column=j).value is not None:
            if availability.cell(row=i,column=j).value is not None:
                available_profs = available_profs + 1
            j = j + 1
        day = c.Day(availability.cell(row=i, column=1).value, available_profs, i - 1)
        i = i + 1
        j = 2
        days.link(day)
    return days


# read all lectures from wishlists and create a sorted linked list
# based on their rating for each course.
def pars_lectures(professors,sessions):
    i = 1
    j = 3
    q = 0
    lectures = c.LinkedList()
    courses = []

    global wish_sheet
    while wish_sheet.cell(row=3, column=i).value is not None:
        while wish_sheet.cell(row=j, column=i).value is not None:
            name = wish_sheet.cell(row=j, column=i).value
            k = 2
            while course_list.cell(row=k ,column=1).value is not name and course_list.cell(row=k ,column=1).value is not None:
                k = k + 1
            lecture = c.Lecture(name, course_list.cell(row=k,column=3).value, course_list.cell(row=k,column=2).value,j-2,course_list.cell(row=k,column=3).value)
            professors[course_list.cell(row=k,column=2).value].add_occurrences(lecture.length)
            lectures.link(lecture)
            q = q + course_list.cell(row=k,column=3).value

            j = j + 1
        days = pars_days()
        if len(courses) is 2:
            days.head = days.head.next.next.next
        courses.append(c.Course(lectures, days,sessions,7))
        lectures = c.LinkedList()
        print("Wishlength:",q)
        q = 0
        i = i + 1
        j = 3

    return courses, professors







